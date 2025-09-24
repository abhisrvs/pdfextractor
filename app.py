from flask import Flask, render_template, request, send_file, flash, redirect, url_for, jsonify, session
import os
import re
import tempfile
from werkzeug.utils import secure_filename
import pdfplumber
import openpyxl
from openpyxl.styles import Font, PatternFill
import io
import zipfile
from datetime import datetime, timedelta
import sqlite3
import requests
import secrets
from functools import wraps
from werkzeug.security import generate_password_hash, check_password_hash

app = Flask(__name__)
app.secret_key = 'your-secret-key-here'  # Change this in production


# --- Absolute DB Paths ---
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
USERS_DB_PATH = os.path.join(BASE_DIR, 'users.db')
HITS_DB_PATH = os.path.join(BASE_DIR, 'hits.db')


# User Database setup
def init_user_db():
    conn = sqlite3.connect(USERS_DB_PATH)
    c = conn.cursor()
    c.execute('''CREATE TABLE IF NOT EXISTS users
                 (id INTEGER PRIMARY KEY AUTOINCREMENT,
                  username TEXT UNIQUE NOT NULL,
                  password TEXT NOT NULL,
                  email TEXT UNIQUE NOT NULL,
                  created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
                  reset_token TEXT,
                  reset_token_expiry DATETIME)''')
    conn.commit()
    conn.close()

# Initialize user database
init_user_db()

def generate_reset_token():
    return secrets.token_urlsafe(32)

# Login required decorator
def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            flash('Please log in to access this page', 'warning')
            return redirect(url_for('login', next=request.url))
        return f(*args, **kwargs)
    return decorated_function


# Database setup
def init_db():
    conn = sqlite3.connect(HITS_DB_PATH)
    c = conn.cursor()
    c.execute('''CREATE TABLE IF NOT EXISTS hits
                 (id INTEGER PRIMARY KEY AUTOINCREMENT,
                  ip TEXT,
                  timestamp DATETIME,
                  endpoint TEXT,
                  country TEXT,
                  city TEXT)''')
    conn.commit()
    conn.close()

# Initialize database
init_db()

# IP Geolocation function
def get_location_info(ip):
    try:
        response = requests.get(f'https://ipapi.co/{ip}/json/')
        if response.status_code == 200:
            data = response.json()
            return {
                'country': data.get('country_name', 'Unknown'),
                'city': data.get('city', 'Unknown')
            }
    except Exception as e:
        print(f"Error getting location: {e}")
    return {'country': 'Unknown', 'city': 'Unknown'}

# Hit counter decorator
def track_hits(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        ip = request.remote_addr
        location = get_location_info(ip)
        conn = sqlite3.connect(HITS_DB_PATH)
        c = conn.cursor()
        c.execute('''INSERT INTO hits (ip, timestamp, endpoint, country, city)
                    VALUES (?, datetime('now'), ?, ?, ?)''',
                  (ip, request.endpoint, location['country'], location['city']))
        conn.commit()
        conn.close()
        return f(*args, **kwargs)
    return decorated_function

# Configuration
UPLOAD_FOLDER = 'uploads'
ALLOWED_EXTENSIONS = {'pdf'}
MAX_FILE_SIZE = 16 * 1024 * 1024  # 16MB

# Ensure upload directory exists
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def extract_emails_from_pdf(pdf_path):
    """Extract emails and employment details (permanent/contract/remote) from PDF file using pdfplumber"""
    email_data = []
    
    try:
        with pdfplumber.open(pdf_path) as pdf:
            for page in pdf.pages:
                text = page.extract_text()
                if text:
                    # Email regex pattern
                    email_pattern = r'\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Z|a-z]{2,}\b'
                    found_emails = re.findall(email_pattern, text)
                    
                    # Process each email
                    for email in found_emails:
                        # Get the context around the email (200 characters before and after for better detection)
                        email_pos = text.find(email)
                        start_pos = max(0, email_pos - 200)
                        end_pos = min(len(text), email_pos + 200)
                        context = text[start_pos:end_pos].lower()
                        
                        # Employment type keywords
                        contract_keywords = ['contract', 'contractor', 'temporary', 'temp', 'fixed term']
                        permanent_keywords = ['permanent', 'full-time', 'full time', 'perm']
                        remote_keywords = ['remote', 'work from home', 'wfh', 'virtual', 'telecommute', 'home-based']
                        
                        # Determine employment type
                        employment_type = 'Unknown'
                        if any(keyword in context for keyword in contract_keywords):
                            employment_type = 'Contract'
                        elif any(keyword in context for keyword in permanent_keywords):
                            employment_type = 'Permanent'
                            
                        # Determine if remote
                        work_location = 'On-site'
                        if any(keyword in context for keyword in remote_keywords):
                            work_location = 'Remote'
                        
                        if email not in [e[0] for e in email_data]:
                            email_data.append((email, employment_type, work_location))
    except Exception as e:
        print(f"Error extracting emails and job types: {e}")
        return []
    
    return email_data

def create_excel_file(email_data, filename="Extracted Emails"):
    """Create Excel file with extracted emails and employment details"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = filename
    
    # Add headers
    headers = ["Email Address", "Domain", "Employment Type", "Work Location", "Status"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    
    # Add email data
    for row, (email, employment_type, work_location) in enumerate(email_data, 2):
        ws.cell(row=row, column=1, value=email)
        # Extract domain
        domain = email.split('@')[1] if '@' in email else ''
        ws.cell(row=row, column=2, value=domain)
        ws.cell(row=row, column=3, value=employment_type)
        ws.cell(row=row, column=4, value=work_location)
        ws.cell(row=row, column=5, value="Valid")
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    return wb

def create_zip_file(file_data):
    """Create a ZIP file containing multiple Excel files"""
    temp_zip = tempfile.NamedTemporaryFile(delete=False, suffix='.zip')
    
    with zipfile.ZipFile(temp_zip.name, 'w', zipfile.ZIP_DEFLATED) as zipf:
        for filename, wb in file_data.items():
            # Create temporary Excel file
            temp_excel = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
            wb.save(temp_excel.name)
            temp_excel.close()
            
            # Add to ZIP
            zipf.write(temp_excel.name, f"{filename}.xlsx")
            
            # Clean up temporary Excel file
            os.unlink(temp_excel.name)
    
    temp_zip.close()
    return temp_zip.name


# Redirect root to /home
@app.route('/')
def root():
    return redirect(url_for('home'))


@app.route('/index')
@login_required
@track_hits
def index():
    return render_template('index.html')

@app.route('/home')
@track_hits
def home():
    return render_template('home.html')

@app.route('/login', methods=['GET', 'POST'])
@track_hits
def login():
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        conn = sqlite3.connect(USERS_DB_PATH)
        c = conn.cursor()
        c.execute('SELECT * FROM users WHERE username = ?', (username,))
        user = c.fetchone()
        conn.close()
        if user and check_password_hash(user[2], password):
            session['user_id'] = user[0]
            session['username'] = user[1]
            flash('Successfully logged in!', 'success')
            next_page = request.args.get('next')
            return redirect(next_page if next_page else url_for('index'))
        flash('Invalid username or password', 'danger')
    return render_template('login.html')

@app.route('/register', methods=['GET', 'POST'])
@track_hits
def register():
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        email = request.form.get('email')
        
        if not all([username, password, email]):
            flash('All fields are required', 'danger')
            return render_template('register.html')
        
        try:
            conn = sqlite3.connect(USERS_DB_PATH)
            c = conn.cursor()
            hashed_password = generate_password_hash(password)
            c.execute('INSERT INTO users (username, password, email) VALUES (?, ?, ?)',
                     (username, hashed_password, email))
            conn.commit()
            conn.close()
            
            flash('Registration successful! Please log in.', 'success')
            return redirect(url_for('login'))
        except sqlite3.IntegrityError:
            flash('Username or email already exists', 'danger')
            return render_template('register.html')
    
    return render_template('register.html')

@app.route('/logout')
def logout():
    session.clear()
    flash('You have been logged out', 'info')
    return redirect(url_for('home'))

@app.route('/profile')
@login_required
def profile():
    conn = sqlite3.connect(USERS_DB_PATH)
    c = conn.cursor()
    c.execute('SELECT username, email, created_at FROM users WHERE id = ?', (session['user_id'],))
    user = c.fetchone()
    conn.close()
    return render_template('profile.html', user=user)

@app.route('/profile/delete', methods=['POST'])
@login_required
def delete_account():
    if request.form.get('confirm_delete') == 'yes':
        conn = sqlite3.connect(USERS_DB_PATH)
        c = conn.cursor()
        c.execute('DELETE FROM users WHERE id = ?', (session['user_id'],))
        conn.commit()
        conn.close()
        session.clear()
        flash('Your account has been permanently deleted.', 'success')
        return redirect(url_for('home'))
    flash('Account deletion requires confirmation.', 'danger')
    return redirect(url_for('profile'))

@app.route('/reset-password', methods=['GET', 'POST'])
def reset_password_request():
    if request.method == 'POST':
        email = request.form.get('email')
        conn = sqlite3.connect(USERS_DB_PATH)
        c = conn.cursor()
        c.execute('SELECT id FROM users WHERE email = ?', (email,))
        user = c.fetchone()
        if user:
            token = generate_reset_token()
            expiry = datetime.now() + timedelta(hours=1)
            c.execute('UPDATE users SET reset_token = ?, reset_token_expiry = ? WHERE email = ?',
                      (token, expiry, email))
            conn.commit()
            # Here you would typically send an email with the reset link
            # For demonstration, we'll just flash the token
            flash(f'Password reset link: {url_for("reset_password", token=token, _external=True)}', 'info')
        else:
            flash('If an account exists with that email, a password reset link will be sent.', 'info')
        conn.close()
        return redirect(url_for('login'))
    return render_template('reset_password_request.html')

@app.route('/reset-password/<token>', methods=['GET', 'POST'])
def reset_password(token):
    conn = sqlite3.connect(USERS_DB_PATH)
    c = conn.cursor()
    c.execute('''SELECT id FROM users 
                 WHERE reset_token = ? AND reset_token_expiry > ?''',
              (token, datetime.now()))
    user = c.fetchone()
    
    if not user:
        conn.close()
        flash('Invalid or expired reset token.', 'danger')
        return redirect(url_for('reset_password_request'))
    
    if request.method == 'POST':
        password = request.form.get('password')
        if password:
            hashed_password = generate_password_hash(password)
            c.execute('''UPDATE users 
                        SET password = ?, reset_token = NULL, reset_token_expiry = NULL 
                        WHERE reset_token = ?''',
                     (hashed_password, token))
            conn.commit()
            flash('Your password has been reset.', 'success')
            return redirect(url_for('login'))
    
    conn.close()
    return render_template('reset_password.html')

@app.route('/profile/update', methods=['POST'])
@login_required
def update_profile():
    username = request.form.get('username')
    email = request.form.get('email')
    current_password = request.form.get('current_password')
    new_password = request.form.get('new_password')
    
    conn = sqlite3.connect(USERS_DB_PATH)
    c = conn.cursor()
    
    try:
        if current_password:
            # Verify current password
            c.execute('SELECT password FROM users WHERE id = ?', (session['user_id'],))
            stored_password = c.fetchone()[0]
            if not check_password_hash(stored_password, current_password):
                flash('Current password is incorrect.', 'danger')
                return redirect(url_for('profile'))
            
            if new_password:
                hashed_password = generate_password_hash(new_password)
                c.execute('UPDATE users SET password = ? WHERE id = ?',
                         (hashed_password, session['user_id']))
        
        if username or email:
            c.execute('UPDATE users SET username = ?, email = ? WHERE id = ?',
                     (username, email, session['user_id']))
            session['username'] = username
        
        conn.commit()
        flash('Profile updated successfully.', 'success')
        
    except sqlite3.IntegrityError:
        flash('Username or email already exists.', 'danger')
    finally:
        conn.close()
    
    return redirect(url_for('profile'))

@app.route('/stats')
@login_required
@track_hits
def stats():
    conn = sqlite3.connect(HITS_DB_PATH)
    c = conn.cursor()
    
    # Get total hits
    c.execute('SELECT COUNT(*) FROM hits')
    total_hits = c.fetchone()[0]
    
    # Get hits by country
    c.execute('''SELECT country, COUNT(*) as count 
                 FROM hits 
                 GROUP BY country 
                 ORDER BY count DESC 
                 LIMIT 10''')
    countries = c.fetchall()
    
    # Get hits by endpoint
    c.execute('''SELECT endpoint, COUNT(*) as count 
                 FROM hits 
                 GROUP BY endpoint 
                 ORDER BY count DESC''')
    endpoints = c.fetchall()
    
    # Get recent hits
    c.execute('''SELECT ip, country, city, timestamp, endpoint 
                 FROM hits 
                 ORDER BY timestamp DESC 
                 LIMIT 10''')
    recent_hits = c.fetchall()
    
    conn.close()
    
    return render_template('stats.html', 
                         total_hits=total_hits,
                         countries=countries,
                         endpoints=endpoints,
                         recent_hits=recent_hits)

@app.route('/upload', methods=['POST'])
@track_hits
def upload_files():
    if 'files' not in request.files:
        return jsonify({'error': 'No files selected'}), 400
    
    files = request.files.getlist('files')
    
    if not files or all(file.filename == '' for file in files):
        return jsonify({'error': 'No files selected'}), 400
    
    valid_files = []
    uploaded_files = []
    
    # Validate and save files
    for file in files:
        if file and allowed_file(file.filename):
            filename = secure_filename(file.filename)
            file_path = os.path.join(UPLOAD_FOLDER, filename)
            file.save(file_path)
            uploaded_files.append(file_path)
            valid_files.append((filename, file_path))
        else:
            return jsonify({'error': f'Invalid file type: {file.filename}. Please upload PDF files only.'}), 400
    
    try:
        results = {}
        total_emails = 0
        
        # Process each file
        for filename, file_path in valid_files:
            email_data = extract_emails_from_pdf(file_path)
            if email_data:
                # Create Excel file for this PDF
                wb = create_excel_file(email_data, filename.replace('.pdf', ''))
                results[filename.replace('.pdf', '')] = wb
                total_emails += len(email_data)
        
        if not results:
            # Clean up uploaded files
            for file_path in uploaded_files:
                if os.path.exists(file_path):
                    os.remove(file_path)
            return jsonify({'error': 'No emails found in any of the PDF files'}), 400
        
        # Create response based on number of files
        if len(results) == 1:
            # Single file - return Excel directly
            filename = list(results.keys())[0]
            wb = list(results.values())[0]
            
            temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
            wb.save(temp_file.name)
            temp_file.close()
            
            # Clean up uploaded files
            for file_path in uploaded_files:
                if os.path.exists(file_path):
                    os.remove(file_path)
            
            return send_file(
                temp_file.name,
                as_attachment=True,
                download_name=f'{filename}_extracted_emails.xlsx',
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
        else:
            # Multiple files - return ZIP
            zip_path = create_zip_file(results)
            
            # Clean up uploaded files
            for file_path in uploaded_files:
                if os.path.exists(file_path):
                    os.remove(file_path)
            
            return send_file(
                zip_path,
                as_attachment=True,
                download_name=f'extracted_emails_{len(results)}_files_{total_emails}_emails.zip',
                mimetype='application/zip'
            )
            
    except Exception as e:
        # Clean up uploaded files on error
        for file_path in uploaded_files:
            if os.path.exists(file_path):
                os.remove(file_path)
        return jsonify({'error': f'Error processing files: {str(e)}'}), 500

@app.route('/delete_temp', methods=['POST'])
def delete_temp_file():
    """Delete temporary file after download"""
    data = request.get_json()
    file_path = data.get('file_path')
    
    if file_path and os.path.exists(file_path):
        try:
            os.remove(file_path)
            return jsonify({'success': True})
        except Exception as e:
            return jsonify({'error': str(e)}), 500
    
    return jsonify({'error': 'File not found'}), 404

if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=8000)
